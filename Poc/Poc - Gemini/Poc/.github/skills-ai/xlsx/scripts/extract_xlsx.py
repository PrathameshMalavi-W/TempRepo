#!/usr/bin/env python3
"""
extract_xlsx.py — Extract structured content from an Excel (.xlsx) file to Markdown.

Usage:
    python extract_xlsx.py <file.xlsx>
    python extract_xlsx.py <file.xlsx> --sheet "Functional Requirements"
    python extract_xlsx.py <file.xlsx> --list-sheets
    python extract_xlsx.py <file.xlsx> --max-rows 100
    python extract_xlsx.py <file.xlsx> --output extracted.md

Requirements:
    pip install openpyxl
"""

import argparse
import sys
from pathlib import Path


def check_dependencies():
    """Check that required libraries are installed."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: openpyxl is not installed.", file=sys.stderr)
        print("Please ask the user for permission, then run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def get_cell_value(cell) -> str:
    """Get cell value as a clean string."""
    if cell.value is None:
        return ""
    value = str(cell.value)
    # Clean up newlines within cells
    value = value.replace("\n", " ").replace("\r", " ").strip()
    return value


def is_row_empty(row) -> bool:
    """Check if all cells in a row are empty."""
    return all(cell.value is None or str(cell.value).strip() == "" for cell in row)


def extract_sheet_to_markdown(ws, max_rows: int | None = None) -> str:
    """Convert an openpyxl worksheet to a markdown table."""
    lines = []
    
    # Collect all rows, skip trailing empty rows
    all_rows = []
    for row in ws.iter_rows():
        if is_row_empty(row):
            continue
        row_values = [get_cell_value(cell) for cell in row]
        # Strip trailing empty cells
        while row_values and row_values[-1] == "":
            row_values.pop()
        if row_values:
            all_rows.append(row_values)

    if not all_rows:
        return "*[Sheet is empty]*"

    # Apply max rows limit
    truncated = False
    if max_rows and len(all_rows) > max_rows + 1:  # +1 for header row
        all_rows = all_rows[:max_rows + 1]
        truncated = True

    # Determine column count (max across all rows)
    col_count = max(len(row) for row in all_rows)

    # Pad all rows to the same column count
    padded_rows = []
    for row in all_rows:
        padded = row + [""] * (col_count - len(row))
        padded_rows.append(padded)

    # Build markdown table
    for i, row in enumerate(padded_rows):
        # Escape pipe characters in cells
        cells = [v.replace("|", "\\|") for v in row]
        lines.append("| " + " | ".join(cells) + " |")
        if i == 0:
            # Add separator after header row
            lines.append("| " + " | ".join(["---"] * col_count) + " |")

    if truncated:
        lines.append("")
        lines.append(f"*[Table truncated at {max_rows} rows. Use --max-rows to increase.]*")

    return "\n".join(lines)


def extract_xlsx(filepath: str, sheet_filter: str | None = None, max_rows: int | None = None) -> str:
    """Extract content from an XLSX file and return as markdown string."""
    import openpyxl

    path = Path(filepath)
    if not path.exists():
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    try:
        # data_only=True evaluates formulas to their cached values
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not open XLSX file: {e}", file=sys.stderr)
        sys.exit(1)

    lines = []
    lines.append(f"# Workbook: {path.name}\n")

    # List all sheets
    sheet_names = wb.sheetnames
    lines.append("## Sheets")
    for i, name in enumerate(sheet_names, start=1):
        lines.append(f"{i}. {name}")
    lines.append("")
    lines.append("---\n")

    # Process sheets
    sheets_to_process = sheet_names
    if sheet_filter:
        # Case-insensitive partial match
        sheets_to_process = [s for s in sheet_names if sheet_filter.lower() in s.lower()]
        if not sheets_to_process:
            lines.append(f"WARNING: No sheet matching '{sheet_filter}' found.")
            lines.append(f"Available sheets: {', '.join(sheet_names)}")
            return "\n".join(lines)

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        lines.append(f"\n## Sheet: {sheet_name}\n")
        table_md = extract_sheet_to_markdown(ws, max_rows)
        lines.append(table_md)
        lines.append("")

    return "\n".join(lines)


def list_sheets(filepath: str):
    """Print all sheet names in the workbook."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        print("Sheets in workbook:")
        for i, name in enumerate(wb.sheetnames, start=1):
            print(f"  {i}. {name}")
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Extract structured content from an XLSX file.")
    parser.add_argument("file", help="Path to the .xlsx file")
    parser.add_argument("--sheet", help="Extract only this sheet (partial name match, case-insensitive)", default=None)
    parser.add_argument("--list-sheets", action="store_true", help="List all sheet names and exit")
    parser.add_argument("--max-rows", type=int, help="Maximum rows to extract per sheet", default=None)
    parser.add_argument("--output", help="Output file path (default: stdout)", default=None)
    args = parser.parse_args()

    check_dependencies()

    if args.list_sheets:
        list_sheets(args.file)
        return

    content = extract_xlsx(args.file, args.sheet, args.max_rows)

    if args.output:
        Path(args.output).write_text(content, encoding="utf-8")
        print(f"Extracted content written to: {args.output}", file=sys.stderr)
    else:
        print(content)


if __name__ == "__main__":
    main()
