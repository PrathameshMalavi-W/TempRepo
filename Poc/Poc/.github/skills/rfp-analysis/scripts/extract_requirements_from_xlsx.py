#!/usr/bin/env python3
"""
Extract requirements from Excel spreadsheet.

Expected Excel format:
    Column A: Requirement ID (e.g., FR-001, NFR-001)
    Column B: Requirement Text/Description
    Column C: Type (FR, NFR, Constraint, etc.)
    Column D: Priority (MUST, SHOULD, COULD)
    Column E (optional): Acceptance Criteria
    Column F (optional): Notes/Tags

Usage:
    python extract_requirements_from_xlsx.py <path_to_file.xlsx> [sheet_name]

Output:
    JSON array of requirement objects
"""

import sys
import json
from pathlib import Path
from typing import List, Dict, Any, Optional

# Try to import openpyxl, with helpful error
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract_requirements(filepath: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Load Excel file and extract requirement rows.
    
    Args:
        filepath: Path to .xlsx file
        sheet_name: Name of sheet to process (defaults to active sheet)
    
    Returns:
        List of requirement dictionaries
    """
    path = Path(filepath)
    
    if not path.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)
    
    if not path.suffix.lower() in ['.xlsx', '.xls']:
        print(f"Error: File must be .xlsx or .xls format, got: {path.suffix}", file=sys.stderr)
        sys.exit(1)
    
    # Load workbook
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Select sheet
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}", file=sys.stderr)
            sys.exit(1)
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    requirements = []
    
    # Skip header row (row 1)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # Extract cell values
        cells = [cell.value for cell in row[:6]]  # Get first 6 columns
        
        req_id = cells[0]
        description = cells[1]
        req_type = cells[2] if len(cells) > 2 else None
        priority = cells[3] if len(cells) > 3 else None
        acceptance_criteria = cells[4] if len(cells) > 4 else None
        notes = cells[5] if len(cells) > 5 else None
        
        # Skip if no ID (empty row)
        if not req_id:
            continue
        
        # Normalize data
        req_id = str(req_id).strip()
        description = str(description).strip() if description else ""
        req_type = str(req_type).strip() if req_type else "FR"
        priority = str(priority).strip().upper() if priority else "SHOULD"
        acceptance_criteria = str(acceptance_criteria).strip() if acceptance_criteria else ""
        notes = str(notes).strip() if notes else ""
        
        # Normalize priority
        if priority not in ['MUST', 'SHOULD', 'COULD']:
            priority = 'SHOULD'  # Default
        
        requirement = {
            'id': req_id,
            'description': description,
            'type': req_type,
            'priority': priority,
            'row': row_idx
        }
        
        if acceptance_criteria:
            requirement['acceptance_criteria'] = acceptance_criteria
        
        if notes:
            requirement['notes'] = notes
        
        requirements.append(requirement)
    
    return requirements


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python extract_requirements_from_xlsx.py <path_to_file.xlsx> [sheet_name]", file=sys.stderr)
        sys.exit(1)
    
    filepath = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    requirements = extract_requirements(filepath, sheet_name)
    
    # Output as JSON
    output = {
        'source_file': filepath,
        'count': len(requirements),
        'requirements': requirements
    }
    
    print(json.dumps(output, indent=2))


if __name__ == '__main__':
    main()
